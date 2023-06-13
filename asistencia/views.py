from django.shortcuts import render
from django.http import FileResponse, HttpResponse
from django.db import connection
from calendar import monthrange
from openpyxl import load_workbook
from .models import Empleado_datos

import win32com
from win32com import client
import os
import pythoncom

def index(request):
    if request.method == 'POST':
        inicio = request.POST.get('inicio')
        titulo = request.POST.get('titulo')
        subtitulo = request.POST.get('subtitulo')
        creado = request.POST.get('fecha')

        fecha_asistencia = inicio.split("/")
        fecha_var = fecha_asistencia[0]
        fecha_var2 = fecha_asistencia[1]

        mesesDic = {
            "01":'Enero',
            "02":'Febrero',
            "03":'Marzo',
            "04":'Abril',
            "05":'Mayo',
            "06":'Junio',
            "07":'Julio ',
            "08":'Agosto',
            "09":'Septiembre',
            "10":'Octubre',
            "11":'Noviembre',
            "12":'Diciembre'
        }
        mes = mesesDic[str(fecha_var)][:9]

        
        # #? Nombre Del Documento
        import xlsxwriter
        workbook = xlsxwriter.Workbook('Asistencia.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.set_column(4, 20, 5)
        worksheet.set_column(3, 0, 12)
        worksheet.set_column(1, 3, 20)
        worksheet.set_column(0, 0, 2)
        worksheet.set_column(3, 3, 7)
        worksheet.set_column(2, 2, 0)
        # #? Orientacion De La Pagina
        worksheet.set_landscape()
        worksheet.protect('Construm@s0000')

        

        # Tamaño De Letra
        font_size = workbook.add_format()
        font_size.set_font_size(8)
        font_size.set_align('center')
        

        # Tamaño De Letra
        font_size2 = workbook.add_format()
        font_size2.set_font_size(7.2)
        font_size2.set_align('left')
        font_size2.set_border(1)
        

        # Tamaño De Letra
        font_size3 = workbook.add_format()
        font_size3.set_font_size(8)
        font_size3.set_align('left')
        font_size3.set_border(1)

        # Tamaño De Letra
        font_size4 = workbook.add_format()
        font_size4.set_font_size(8)
        font_size4.set_align('center')
        font_size4.set_border(1)

        
        # Datos Incrustados En El XLSX
        worksheet.write('J1',titulo.upper(),font_size)
        worksheet.write('J2','PERIODO DEL 11 AL 25 DE '+ mes.upper() + ' DEL AÑO 20'+ fecha_var2 ,font_size)
        worksheet.write('B2',subtitulo ,font_size)
        worksheet.write('R1','SISTEMA v1.0' ,font_size)
        worksheet.write('R2','CREADO EL: '+creado ,font_size)



        worksheet.write('B66','_____________',font_size)
        worksheet.write('B67','ELABORADO' ,font_size)
        worksheet.write('B68','IRIS MENDOZA',font_size)
        
        worksheet.write('I66','_________________',font_size)
        worksheet.write('I67','ING. TEDDI GARCIA ' ,font_size)
        worksheet.write('I68','GERENTE DE PRODUCCIÓN: ' ,font_size)


        worksheet.write('R66','_____________________',font_size)
        worksheet.write('R67','LICDA. HILDA VALENCIA' ,font_size)
        worksheet.write('R68','JEFE DE ADMINISTRACIÓN' ,font_size)

        altura_titulo = '4'
        altura_columna = 4

        def empleado_function():
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT DISTINCT numero_empleado, nombre, estatus,estado_empleado FROM asistencia_empleados
                    INNER JOIN asistencia_empleado_datos ON asistencia_empleado_datos.id = asistencia_empleados.nombre_empleado_id
                    WHERE NOT estado_empleado = 0
                """)
                empleado = cursor.fetchall()
            worksheet.write('A'+altura_titulo, 'N°: ',font_size3)
            id_nombre = [item[0] for item in empleado]
            worksheet.write_column(altura_columna, 0, id_nombre,font_size2)

            worksheet.write('B'+altura_titulo, 'NOMBRE:',font_size3)
            id_nombre = [item[1] for item in empleado]
            worksheet.write_column(altura_columna, 1, id_nombre,font_size2)

            worksheet.write('D'+altura_titulo, 'ESTATUS:',font_size3)
            id_nombre = [item[2] for item in empleado]
            worksheet.write_column(altura_columna, 3, id_nombre,font_size2)
        empleado_function()

        def all_fechas():

            def buscador_fecha(dia = '0',fecha = '0'):
                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        SELECT titulo_asistencia, fecha,estados FROM asistencia_asistencia_empleado
                        LEFT JOIN asistencia_asistencia ON asistencia_asistencia.id = asistencia_asistencia_empleado.asistencia_id
                        LEFT JOIN asistencia_empleados ON asistencia_empleados.id = asistencia_asistencia_empleado.empleados_id
						INNER JOIN asistencia_empleado_datos ON asistencia_empleado_datos.id = asistencia_empleados.nombre_empleado_id
                        WHERE NOT estado_empleado = 0 AND fecha = '{dia}/{fecha}' ORDER BY numero_empleado
                    """)
                    asistencia11 = cursor.fetchall()
                return asistencia11
            
            def fecha11():
                asistencia11 = buscador_fecha('11',inicio)

                fecha = [item[0] == 'DIA - 11' for item in asistencia11]

                if fecha:
                    worksheet.write('E'+altura_titulo, 'DIA 11:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 4, nombre,font_size4)
            fecha11()

            def fecha12():
                asistencia11 = buscador_fecha('12',inicio)

                fecha = [item[0] == 'DIA - 12' for item in asistencia11]
                if fecha:
                    worksheet.write('F'+altura_titulo, 'DIA 12:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 5, nombre,font_size4)
            fecha12()

            def fecha13():
                asistencia11 = buscador_fecha('13',inicio)

                fecha = [item[0] == 'DIA - 13' for item in asistencia11]

                if fecha:
                    worksheet.write('G'+altura_titulo, 'DIA 13:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 6, nombre,font_size4)
            fecha13()

            def fecha14():
                asistencia11 = buscador_fecha('14',inicio)

                fecha = [item[0] == 'DIA - 14' for item in asistencia11]
                if fecha:
                    worksheet.write('H'+altura_titulo, 'DIA 14:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 7, nombre,font_size4)
            fecha14()

            def fecha15():
                asistencia11 = buscador_fecha('15',inicio)

                fecha = [item[0] == 'DIA - 15' for item in asistencia11]
                if fecha:
                    worksheet.write('I'+altura_titulo, 'DIA 15:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 8, nombre,font_size4)
            fecha15()

            def fecha16():
                asistencia11 = buscador_fecha('16',inicio)       

                fecha = [item[0] == 'DIA - 16' for item in asistencia11]
                if fecha:
                    worksheet.write('J'+altura_titulo, 'DIA 16:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 9, nombre,font_size4)
            fecha16()

            def fecha17():
                asistencia11 = buscador_fecha('17',inicio)

                fecha = [item[0] == 'DIA - 17' for item in asistencia11]
                if fecha:
                    worksheet.write('K'+altura_titulo, 'DIA 17:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 10, nombre,font_size4)
            fecha17()

            def fecha18():
                asistencia11 = buscador_fecha('18',inicio)

                fecha = [item[0] == 'DIA - 18' for item in asistencia11]
                if fecha:
                    worksheet.write('L'+altura_titulo, 'DIA 18:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 11, nombre,font_size4)
            fecha18()

            def fecha19():
                asistencia11 = buscador_fecha('19',inicio)

                fecha = [item[0] == 'DIA - 19' for item in asistencia11]
                if fecha:
                    worksheet.write('M'+altura_titulo, 'DIA 19:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 12, nombre,font_size4)
            fecha19()

            def fecha20():
                asistencia11 = buscador_fecha('20',inicio)

                fecha = [item[0] == 'DIA - 20' for item in asistencia11]
                if fecha:
                    worksheet.write('N'+altura_titulo, 'DIA 20:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 13, nombre,font_size4)
            fecha20()

            def fecha21():
                asistencia11 = buscador_fecha('21',inicio)

                fecha = [item[0] == 'DIA - 21' for item in asistencia11]
                if fecha:
                    worksheet.write('O'+altura_titulo, 'DIA 21:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 14, nombre,font_size4)
            fecha21()

            def fecha22():
                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        SELECT titulo_asistencia, fecha,estados FROM asistencia_asistencia_empleado
                        LEFT JOIN asistencia_asistencia ON asistencia_asistencia.id = asistencia_asistencia_empleado.asistencia_id
                        LEFT JOIN asistencia_empleados ON asistencia_empleados.id = asistencia_asistencia_empleado.empleados_id
						INNER JOIN asistencia_empleado_datos ON asistencia_empleado_datos.id = asistencia_empleados.nombre_empleado_id
                        WHERE NOT estado_empleado = 0 AND fecha = '22/{inicio}' ORDER BY numero_empleado
                    """)
                    asistencia11 = cursor.fetchall()

                fecha = [item[0] == 'DIA - 22' for item in asistencia11]
                if fecha:
                    worksheet.write('P'+altura_titulo, 'DIA 22:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 15, nombre,font_size4)
            fecha22()

            def fecha23():
                asistencia11 = buscador_fecha('23',inicio)

                fecha = [item[0] == 'DIA - 23' for item in asistencia11]
                if fecha:
                    worksheet.write('Q'+altura_titulo, 'DIA 23:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 16, nombre,font_size4)
            fecha23()

            def fecha24():
                asistencia11 = buscador_fecha('24',inicio)

                fecha = [item[0] == 'DIA - 24' for item in asistencia11]
                if fecha:
                    worksheet.write('R'+altura_titulo, 'DIA 24:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 17, nombre,font_size4)
            fecha24()

            def fecha25():
                asistencia11 = buscador_fecha('25',inicio)

                fecha = [item[0] == 'DIA - 25' for item in asistencia11]
                if fecha:
                    worksheet.write('S'+altura_titulo, 'DIA 25:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 18, nombre,font_size4)
            fecha25()

        all_fechas()
        workbook.close()

        def create_pdf():
            import win32com
            from win32com import client
            import os
            import pythoncom
            #currentDir = os.path.abspath('.')
            currentDir = os.getcwd()
            xlApp = win32com.client.Dispatch("Excel.Application",pythoncom.CoInitialize())
            books = xlApp.Workbooks.Open(os.path.join(currentDir,"Asistencia.xlsx"))    
            ws = books.Worksheets[0]
            ws.Visible = 1
            ws.ExportAsFixedFormat(0,os.path.join(currentDir,"Asistencia"))
            books.Close()
        create_pdf()
        return FileResponse(open('Asistencia.pdf', 'rb'), content_type='application/pdf')
    return render(request,'index.html')

def index2(request):
    if request.method == 'POST':
        inicio = request.POST.get('inicio')
        titulo = request.POST.get('titulo')
        subtitulo = request.POST.get('subtitulo')
        creado = request.POST.get('fecha')

        fecha_asistencia = inicio.split("/")
        fecha_var = fecha_asistencia[0]
        fecha_var2 = fecha_asistencia[1]

        mesesDic = {
            "01":'Enero',
            "02":'Febrero',
            "03":'Marzo',
            "04":'Abril',
            "05":'Mayo',
            "06":'Junio',
            "07":'Julio ',
            "08":'Agosto',
            "09":'Septiembre',
            "10":'Octubre',
            "11":'Noviembre',
            "12":'Diciembre'
        }
        mes = mesesDic[str(fecha_var)][:9]

        
        # #? Nombre Del Documento
        import xlsxwriter
        workbook = xlsxwriter.Workbook('Asistencia.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.set_column(4, 20, 5)
        worksheet.set_column(3, 0, 12)
        worksheet.set_column(1, 3, 20)
        worksheet.set_column(0, 0, 2)
        worksheet.set_column(3, 3, 7)
        worksheet.set_column(2, 2, 0)
        # #? Orientacion De La Pagina
        worksheet.set_landscape()
        worksheet.protect('Construm@s0000')

        

        # Tamaño De Letra
        font_size = workbook.add_format()
        font_size.set_font_size(8)
        font_size.set_align('center')
        

        # Tamaño De Letra
        font_size2 = workbook.add_format()
        font_size2.set_font_size(7.2)
        font_size2.set_align('left')
        font_size2.set_border(1)
        

        # Tamaño De Letra
        font_size3 = workbook.add_format()
        font_size3.set_font_size(8)
        font_size3.set_align('left')
        font_size3.set_border(1)

        # Tamaño De Letra
        font_size4 = workbook.add_format()
        font_size4.set_font_size(8)
        font_size4.set_align('center')
        font_size4.set_border(1)

        
        # Datos Incrustados En El XLSX
        worksheet.write('J1',titulo.upper(),font_size)
        worksheet.write('J2','PERIODO DEL 26 AL 10 DE '+ mes.upper() + ' DEL AÑO 20'+ fecha_var2 ,font_size)
        worksheet.write('B2',subtitulo ,font_size)
        worksheet.write('R1','SISTEMA v1.0' ,font_size)
        worksheet.write('R2','CREADO EL: '+creado ,font_size)


        worksheet.write('B66','_____________',font_size)
        worksheet.write('B67','ELABORADO' ,font_size)
        worksheet.write('B68','IRIS MENDOZA',font_size)
        
        worksheet.write('I66','_________________',font_size)
        worksheet.write('I67','ING. TEDDI GARCIA ' ,font_size)
        worksheet.write('I68','GERENTE DE PRODUCCIÓN: ' ,font_size)


        worksheet.write('R66','_____________________',font_size)
        worksheet.write('R67','LICDA. HILDA VALENCIA' ,font_size)
        worksheet.write('R68','JEFE DE ADMINISTRACIÓN' ,font_size)

        altura_titulo = '4'
        altura_columna = 4

        def empleado_function():
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT DISTINCT numero_empleado, nombre, estatus,estado_empleado FROM asistencia_empleados
                    INNER JOIN asistencia_empleado_datos ON asistencia_empleado_datos.id = asistencia_empleados.nombre_empleado_id
                    WHERE NOT estado_empleado = 0
                """)
                empleado = cursor.fetchall()
            worksheet.write('A'+altura_titulo, 'N°: ',font_size3)
            id_nombre = [item[0] for item in empleado]
            worksheet.write_column(altura_columna, 0, id_nombre,font_size2)

            worksheet.write('B'+altura_titulo, 'NOMBRE:',font_size3)
            id_nombre = [item[1] for item in empleado]
            worksheet.write_column(altura_columna, 1, id_nombre,font_size2)

            worksheet.write('D'+altura_titulo, 'ESTATUS:',font_size3)
            id_nombre = [item[2] for item in empleado]
            worksheet.write_column(altura_columna, 3, id_nombre,font_size2)
        empleado_function()

        def all_fechas():

            def help_fecha():
                day_fecha = fecha_var.split('0')[1]
                num_days = monthrange(20+int(fecha_var2), int(day_fecha))[1] # num_days = 31
                return num_days

            def buscador_fecha(dia = '0',fecha = '0'):
                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        SELECT titulo_asistencia2, fecha2,estados FROM asistencia_asistencia2_empleado2
                        LEFT JOIN asistencia_asistencia2 ON asistencia_asistencia2.id = asistencia_asistencia2_empleado2.asistencia2_id
                        LEFT JOIN asistencia_empleados ON asistencia_empleados.id = asistencia_asistencia2_empleado2.empleados_id
						INNER JOIN asistencia_empleado_datos ON asistencia_empleado_datos.id = asistencia_empleados.nombre_empleado_id
                        WHERE NOT estado_empleado = 0 AND fecha2 = '{dia}/{fecha}' ORDER BY numero_empleado
                    """)
                    asistencia11 = cursor.fetchall()
                return asistencia11
                
            def fecha_relativa():
                fecha_suma =int(fecha_var)+1
                return '0'+str(fecha_suma)+'/'+fecha_var2
            
            def fecha26():
                asistencia11 = buscador_fecha('26',inicio)
                fecha = [item[0] == 'DIA - 26' for item in asistencia11]

                if fecha:
                    worksheet.write('E'+altura_titulo, 'DIA 26:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 4, nombre,font_size4)
            fecha26()

            def fecha27():
                asistencia11 = buscador_fecha('27',inicio)
                fecha = [item[0] == 'DIA - 27' for item in asistencia11]
                if fecha:
                    worksheet.write('F'+altura_titulo, 'DIA 27:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 5, nombre,font_size4)
            fecha27()

            def fecha28():
                asistencia11 = buscador_fecha('28',inicio)

                fecha = [item[0] == 'DIA - 28' for item in asistencia11]

                if fecha:
                    worksheet.write('G'+altura_titulo, 'DIA 28:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 6, nombre,font_size4)
            fecha28()

            def fecha29():
                asistencia11 = buscador_fecha('29',inicio)

                fecha = [item[0] == 'DIA - 29' for item in asistencia11]
                if fecha:
                    worksheet.write('H'+altura_titulo, 'DIA 29:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 7, nombre,font_size4)
            fecha29()

            def fecha30():
                asistencia11 = buscador_fecha('30',inicio)

                fecha = [item[0] == 'DIA - 30' for item in asistencia11]
                if fecha:
                    worksheet.write('I'+altura_titulo, 'DIA 30:',font_size4)
                    nombre = [item[2] for item in asistencia11]
                    worksheet.write_column(altura_columna, 8, nombre,font_size4)
            fecha30()

            fecha_dia_mes = help_fecha()
            if fecha_dia_mes == 30:
                def fecha_con_30():
                    def fecha1():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('1',inicio_fecha)

                        fecha = [item[0] == 'DIA - 1' for item in asistencia11]
                        if fecha:
                            worksheet.write('J'+altura_titulo, 'DIA 1:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 9, nombre,font_size4)
                    fecha1()

                    def fecha2():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('2',inicio_fecha)

                        fecha = [item[0] == 'DIA - 2' for item in asistencia11]
                        if fecha:
                            worksheet.write('K'+altura_titulo, 'DIA 2:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 10, nombre,font_size4)
                    fecha2()

                    def fecha3():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('3',inicio_fecha)

                        fecha = [item[0] == 'DIA - 3' for item in asistencia11]
                        if fecha:
                            worksheet.write('L'+altura_titulo, 'DIA 3:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 11, nombre,font_size4)
                    fecha3()

                    def fecha4():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('4',inicio_fecha)

                        fecha = [item[0] == 'DIA - 4' for item in asistencia11]
                        if fecha:
                            worksheet.write('M'+altura_titulo, 'DIA 4:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 12, nombre,font_size4)
                    fecha4()

                    def fecha5():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('5',inicio_fecha)

                        fecha = [item[0] == 'DIA - 5' for item in asistencia11]
                        if fecha:
                            worksheet.write('N'+altura_titulo, 'DIA 5:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 13, nombre,font_size4)
                    fecha5()

                    def fecha6():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('6',inicio_fecha)

                        fecha = [item[0] == 'DIA - 6' for item in asistencia11]
                        if fecha:
                            worksheet.write('O'+altura_titulo, 'DIA 6:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 14, nombre,font_size4)
                    fecha6()

                    def fecha7():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('7',inicio_fecha)

                        fecha = [item[0] == 'DIA - 7' for item in asistencia11]
                        if fecha:
                            worksheet.write('P'+altura_titulo, 'DIA 7:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 15, nombre,font_size4)
                    fecha7()

                    def fecha8():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('8',inicio_fecha)

                        fecha = [item[0] == 'DIA - 8' for item in asistencia11]
                        if fecha:
                            worksheet.write('Q'+altura_titulo, 'DIA 8:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 16, nombre,font_size4)
                    fecha8()

                    def fecha9():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('9',inicio_fecha)

                        fecha = [item[0] == 'DIA - 9' for item in asistencia11]
                        if fecha:
                            worksheet.write('R'+altura_titulo, 'DIA 9:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 17, nombre,font_size4)
                    fecha9()

                    def fecha10():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('10',inicio_fecha)
                        
                        fecha = [item[0] == 'DIA - 10' for item in asistencia11]
                        if fecha:
                            worksheet.write('S'+altura_titulo, 'DIA 10:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 18, nombre,font_size4)
                    fecha10()
                fecha_con_30()
            else:
                def fecha_con_31():

                    def fecha31():
                        asistencia11 = buscador_fecha('31',inicio)

                        fecha = [item[0] == 'DIA - 31' for item in asistencia11]
                        if fecha:
                            worksheet.write('J'+altura_titulo, 'DIA 31:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 9, nombre,font_size4)
                    fecha31()

                    def fecha1():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('1',inicio_fecha)

                        fecha = [item[0] == 'DIA - 1' for item in asistencia11]
                        if fecha:
                            worksheet.write('K'+altura_titulo, 'DIA 1:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 10, nombre,font_size4)
                    fecha1()

                    def fecha2():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('2',inicio_fecha)

                        fecha = [item[0] == 'DIA - 2' for item in asistencia11]
                        if fecha:
                            worksheet.write('L'+altura_titulo, 'DIA 2:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 11, nombre,font_size4)
                    fecha2()

                    def fecha3():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('3',inicio_fecha)

                        fecha = [item[0] == 'DIA - 3' for item in asistencia11]
                        if fecha:
                            worksheet.write('M'+altura_titulo, 'DIA 3:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 12, nombre,font_size4)
                    fecha3()

                    def fecha4():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('4',inicio_fecha)
                        fecha = [item[0] == 'DIA - 4' for item in asistencia11]
                        if fecha:
                            worksheet.write('N'+altura_titulo, 'DIA 4:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 13, nombre,font_size4)
                    fecha4()

                    def fecha5():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('5',inicio_fecha)

                        fecha = [item[0] == 'DIA - 5' for item in asistencia11]
                        if fecha:
                            worksheet.write('O'+altura_titulo, 'DIA 5:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 14, nombre,font_size4)
                    fecha5()

                    def fecha6():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('6',inicio_fecha)

                        fecha = [item[0] == 'DIA - 6' for item in asistencia11]
                        if fecha:
                            worksheet.write('P'+altura_titulo, 'DIA 6:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 15, nombre,font_size4)
                    fecha6()

                    def fecha7():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('7',inicio_fecha)

                        fecha = [item[0] == 'DIA - 7' for item in asistencia11]
                        if fecha:
                            worksheet.write('Q'+altura_titulo, 'DIA 7:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 16, nombre,font_size4)
                    fecha7()

                    def fecha8():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('8',inicio_fecha)

                        fecha = [item[0] == 'DIA - 8' for item in asistencia11]
                        if fecha:
                            worksheet.write('R'+altura_titulo, 'DIA 8:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 17, nombre,font_size4)
                    fecha8()

                    def fecha9():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('9',inicio_fecha)

                        fecha = [item[0] == 'DIA - 9' for item in asistencia11]
                        if fecha:
                            worksheet.write('S'+altura_titulo, 'DIA 9:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 18, nombre,font_size4)
                    fecha9()

                    def fecha10():
                        inicio_fecha = fecha_relativa()
                        asistencia11 = buscador_fecha('10',inicio_fecha)
                        fecha = [item[0] == 'DIA - 10' for item in asistencia11]
                        if fecha:
                            worksheet.write('S'+altura_titulo, 'DIA 10:',font_size4)
                            nombre = [item[2] for item in asistencia11]
                            worksheet.write_column(altura_columna, 18, nombre,font_size4)
                    fecha10()
                fecha_con_31()
        all_fechas()
        workbook.close()

        def create_pdf():
            #currentDir = os.path.abspath('.')
            currentDir = os.getcwd()
            xlApp = win32com.client.Dispatch("Excel.Application",pythoncom.CoInitialize())
            books = xlApp.Workbooks.Open(os.path.join(currentDir,"Asistencia.xlsx"))    
            ws = books.Worksheets[0]
            ws.Visible = 1
            ws.ExportAsFixedFormat(0,os.path.join(currentDir,"Asistencia"))
            books.Close()
        create_pdf()
        return FileResponse(open('Asistencia.pdf', 'rb'), content_type='application/pdf')
    return render(request,'index2.html')

def ap_personal(request):
    if request.method == 'POST':
        empleado = request.POST.get('empleado')
        accion = request.POST.get('accion')
        fecha = request.POST.get('fecha')
        observacion = request.POST.get('observacion')

        def fecha_es():
            fecha_asistencia = fecha.split("-")
            fecha_var = fecha_asistencia[2]
            fecha_var2 = fecha_asistencia[1]
            fecha_var3 = fecha_asistencia[0]
            mesesDic = {
                "01":'Enero',
                "02":'Febrero',
                "03":'Marzo',
                "04":'Abril',
                "05":'Mayo',
                "06":'Junio',
                "07":'Julio ',
                "08":'Agosto',
                "09":'Septiembre',
                "10":'Octubre',
                "11":'Noviembre',
                "12":'Diciembre'
            }
            mes = mesesDic[str(fecha_var2)][:9]
            fecha_final = f"{fecha_var} de {mes} del {fecha_var3}"
            return fecha_final

        empleado = empleado.split('|')[0]
        with connection.cursor() as cursor:
            cursor.execute("SELECT nombre_empleado_dato, cargo, departamento, honorario, dia_ingreso, numero_empleado FROM asistencia_empleado_datos WHERE nombre_empleado_dato = %s",[empleado.strip()])
            empleado_dato = cursor.fetchall()
        
        for datos_empleado in empleado_dato:
            def crear_ap():
                currentDir = os.getcwd()
                workbook = load_workbook(os.path.join(currentDir,"Ap.xlsx"))
                worksheet = workbook.active

                worksheet['G7'] = datos_empleado[0]
                worksheet['G8'] = datos_empleado[1]
                worksheet['G9'] = datos_empleado[2]
                worksheet['Q8'] = datos_empleado[3]
                worksheet['Q9'] = datos_empleado[4]
                worksheet['Q10'] = datos_empleado[5]
                worksheet['C44'] = fecha_es()
                worksheet['B49'] = observacion

                #! Primera Seccion

                worksheet['D14'] = 'X' if accion == '1' else ''
                worksheet['D16'] = 'X' if accion == '2' else ''
                worksheet['D18'] = 'X' if accion == '3' else ''
                worksheet['D20'] = 'X' if accion == '4' else ''

                worksheet['D22'] = 'X' if accion == '5' else ''
                worksheet['D24'] = 'X' if accion == '6' else ''
                worksheet['D26'] = 'X' if accion == '7' else ''
                worksheet['D28'] = 'X' if accion == '8' else ''

                worksheet['D30'] = 'X' if accion == '9' else ''
                worksheet['D32'] = 'X' if accion == '10' else ''
                worksheet['D34'] = 'X' if accion == '11' else ''
                worksheet['D36'] = 'X' if accion == '12' else ''
                worksheet['D38'] = 'X' if accion == '13' else ''
                worksheet['D40'] = 'X' if accion == '14' else ''

                #! Segunda Seccion

                worksheet['L14'] = 'X' if accion == '15' else ''
                worksheet['L16'] = 'X' if accion == '16' else ''
                worksheet['L18'] = 'X' if accion == '17' else ''
                worksheet['L20'] = 'X' if accion == '18' else ''

                worksheet['L22'] = 'X' if accion == '19' else ''
                worksheet['L24'] = 'X' if accion == '20' else ''
                worksheet['L26'] = 'X' if accion == '21' else ''
                worksheet['L28'] = 'X' if accion == '22' else ''

                worksheet['L30'] = 'X' if accion == '23' else ''
                worksheet['L32'] = 'X' if accion == '24' else ''
                worksheet['L34'] = 'X' if accion == '25' else ''
                worksheet['L36'] = 'X' if accion == '26' else ''
                worksheet['L38'] = 'X' if accion == '27' else ''
                worksheet['L40'] = 'X' if accion == '28' else ''

                workbook.save("Ap.xlsx")
            crear_ap()
        
        return FileResponse(open('Ap.xlsx', 'rb'), content_type='application/xlsx')

    empleados = Empleado_datos.objects.all()
    return render(request, 'ap_personal.html',{'empleados': empleados})